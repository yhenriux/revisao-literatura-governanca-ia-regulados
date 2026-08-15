#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MetaGrade Python + LLM Adjudicator

Pipeline para Revisão Sistemática de Literatura com apoio combinado de Python e LLM.
Foco: Governança Conversacional em sistemas baseados em LLMs em ambientes regulados.

O que faz:
1. Lê PDFs de uma pasta ou um PDF único.
2. Extrai texto com PyMuPDF.
3. Remove referências da análise temática principal.
4. Faz triagem determinística em Python por taxonomia controlada.
5. Seleciona evidências literais por página e seção provável.
6. Usa LLM para automatizar o que exigiria leitura humana:
   - correção bibliográfica;
   - elegibilidade PRISMA;
   - classificação metodológica;
   - appraisal CASP/JBI adaptado;
   - CERQual;
   - extração temática;
   - síntese da contribuição;
   - decisão final assistida por LLM.
7. Valida se as evidências citadas pela LLM existem no texto extraído.
8. Gera checkpoints JSONL durante a execução.
9. Exporta um único XLSX com múltiplas abas.

Instalação:
    pip install pandas openpyxl pymupdf tqdm

Uso padrão com Cursor CLI / Composer 2.5:
    python metagrade_python_llm_adjudicator_cursor.py --provider cursor-cli --model composer-2.5 --no-progressbar

Uso com API OpenAI-compatible:
    setx LLM_API_KEY "SUA_CHAVE"
    python metagrade_python_llm_adjudicator_cursor.py --input "C:\\...\\pdfs" --output "C:\\...\\metagrade_llm_output" --provider openai --model gpt-4o-mini

Uso com OpenRouter ou outro endpoint OpenAI-compatible:
    setx LLM_API_KEY "SUA_CHAVE"
    python metagrade_python_llm_adjudicator.py --provider openai --api-base "https://openrouter.ai/api/v1" --model "anthropic/claude-3.5-sonnet"

Uso com Ollama local, se quiser:
    python metagrade_python_llm_adjudicator.py --provider ollama --api-base "http://localhost:11434" --model llama3.2:3b

Teste com 1 PDF:
    python metagrade_python_llm_adjudicator.py --input "C:\\Users\\yhenr\\Downloads\\metagrade_teste_1pdf" --output "C:\\Users\\yhenr\\Downloads\\metagrade_llm_teste" --limit 1

Nota metodológica:
Este script automatiza adjudicação assistida por LLM. Para publicação, descreva como
"triagem e avaliação assistidas por LLM com validação determinística de evidências e auditoria".
Não descreva como julgamento humano independente.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import random
import re
import time
import traceback
import urllib.error
import urllib.request
import subprocess
import shutil
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd

try:
    import fitz  # PyMuPDF
except ImportError as exc:
    raise SystemExit("Instale PyMuPDF: pip install pymupdf") from exc

try:
    from tqdm import tqdm
except Exception:
    tqdm = None


# ============================================================
# CONFIGURAÇÃO
# ============================================================

DEFAULT_INPUT_DIR = r"C:\Users\yhenr\Downloads\artigos_unificados\revisao_literatura_aigovernanca\fulltext_repository\pdfs"
DEFAULT_OUTPUT_DIR = r"C:\Users\yhenr\Downloads\artigos_unificados\revisao_literatura_aigovernanca\metagrade_llm_output"
SCRIPT_VERSION = "python_llm_adjudicator_cursor_cli_v1.3_failfast_agent_default"

PROJECT_SCOPE = """
Revisão sistemática de literatura sobre Governança Conversacional em sistemas baseados em Modelos de Linguagem de Grande Escala (LLMs) aplicados a ambientes regulados.
Interessa especialmente: governança de IA generativa, sistemas conversacionais, chatbots, agentes, RAG, guardrails, observabilidade, auditoria, logs, tracing, controle de risco, compliance, supervisão humana, accountability, handoff, reparo, contestação, monitoramento pós-implantação, aprendizagem operacional, governança de conhecimento, privacidade, segurança, vieses, explicabilidade e contextos regulados como saúde, financeiro, jurídico, setor público, educação regulada e seguros.
""".strip()

OUTPUT_COLUMNS = [
    "record_id", "study_id", "run_id", "script_version", "created_at",
    "raw_file_name", "pdf_path", "pdf_hash", "text_hash", "page_count",
    "text_extraction_status", "text_char_count", "body_char_count_no_references",
    "possible_scanned_pdf", "references_detected", "tables_detected", "figures_detected",
    "extracted_text_path", "llm_raw_output_path",

    "py_title", "py_authors", "py_year", "py_venue", "py_doi", "py_publication_type", "py_language",
    "py_scope_total_score", "py_scope_grade", "py_triage_decision", "py_triage_rationale",
    "py_matched_terms_summary", "py_top_evidence_excerpt", "py_top_evidence_page", "py_top_evidence_section",

    "llm_title", "llm_authors", "llm_year", "llm_venue", "llm_doi", "llm_publication_type", "llm_peer_reviewed", "llm_citation_apa",
    "llm_research_type", "llm_method_family", "llm_study_design", "llm_sample_description", "llm_sample_size",
    "llm_sector", "llm_regulated_domain", "llm_llm_specificity", "llm_technology_context",

    "llm_title_abstract_decision", "llm_full_text_decision", "llm_final_decision", "llm_prisma_stage",
    "llm_inclusion_codes", "llm_exclusion_codes", "llm_exclusion_reason", "llm_decision_rationale",
    "llm_decision_confidence", "llm_manual_spotcheck_required", "llm_manual_spotcheck_reason",

    "appraisal_objective_clarity", "appraisal_method_fit", "appraisal_sampling_adequacy",
    "appraisal_data_collection_transparency", "appraisal_analysis_rigor", "appraisal_ethics_reflexivity",
    "appraisal_results_clarity", "appraisal_limitations_transparency", "appraisal_evidence_support", "appraisal_transferability",
    "appraisal_total_score", "appraisal_quality_grade", "appraisal_notes",

    "cerqual_methodological_limitations", "cerqual_coherence", "cerqual_adequacy", "cerqual_relevance",
    "cerqual_overall_confidence", "cerqual_explanation",

    "coding_open_codes", "coding_axial_codes", "coding_theme", "coding_subtheme", "coding_concepts",
    "coding_model_layers", "coding_rq_alignment", "coding_confidence",

    "synthesis_main_claim", "synthesis_contribution", "synthesis_limitations", "synthesis_use_in_review",
    "llm_evidence_json", "validated_evidence_count", "invalid_evidence_count", "invalid_evidence_notes",

    "consensus_final_decision", "consensus_score", "consensus_grade", "consensus_priority", "consensus_rationale",
    "data_quality_flags", "audit_status", "audit_notes", "notes_for_reproducibility",
]


TAXONOMY: Dict[str, Dict[str, Any]] = {
    "llm_genai": {
        "label": "LLM e IA generativa",
        "weight": 18,
        "strong": [
            "large language model", "large language models", "llm", "llms", "generative ai",
            "generative artificial intelligence", "foundation model", "foundation models", "chatgpt", "gpt-3", "gpt-4",
            "modelo de linguagem de grande escala", "modelos de linguagem de grande escala", "ia generativa",
            "inteligência artificial generativa", "inteligencia artificial generativa", "modelo fundacional", "modelos fundacionais",
        ],
        "weak": ["natural language processing", "nlp", "artificial intelligence", "machine learning", "deep learning", "language model", "text generation", "processamento de linguagem natural", "inteligência artificial", "inteligencia artificial", "aprendizado de máquina", "modelo de linguagem"],
    },
    "conversational_ai": {
        "label": "IA conversacional e agentes",
        "weight": 16,
        "strong": ["chatbot", "chatbots", "conversational ai", "conversational agent", "conversational agents", "dialogue system", "dialog system", "virtual assistant", "virtual assistants", "digital assistant", "voice assistant", "ia conversacional", "agente conversacional", "agentes conversacionais", "assistente virtual", "assistentes virtuais"],
        "weak": ["conversation", "dialogue", "interaction", "automated interaction", "human-computer interaction", "conversa", "diálogo", "dialogo", "interação", "interacao"],
    },
    "governance_accountability": {
        "label": "Governança e accountability",
        "weight": 20,
        "strong": ["governance", "accountability", "algorithmic accountability", "ai governance", "responsible ai", "oversight", "transparency", "explainability", "interpretability", "responsibility", "governança", "governanca", "responsabilização", "responsabilizacao", "prestação de contas", "prestacao de contas", "ia responsável", "ia responsavel", "supervisão", "supervisao", "transparência", "transparencia", "explicabilidade", "interpretabilidade"],
        "weak": ["trust", "ethics", "fairness", "bias", "legitimacy", "public values", "confiança", "confianca", "ética", "etica", "justiça", "justica", "vieses", "viés", "vies", "legitimidade"],
    },
    "regulated_high_stakes": {
        "label": "Ambientes regulados e alto risco",
        "weight": 14,
        "strong": ["regulated", "regulatory", "high-stakes", "high stakes", "public sector", "government", "healthcare", "health care", "financial services", "banking", "insurance", "legal", "criminal justice", "education", "public administration", "regulado", "regulatória", "regulatoria", "alto risco", "setor público", "setor publico", "governo", "saúde", "saude", "serviços financeiros", "servicos financeiros", "bancário", "bancario", "seguros", "jurídico", "juridico", "justiça criminal", "justica criminal", "educação", "educacao", "administração pública", "administracao publica"],
        "weak": ["compliance", "law", "policy", "standard", "institutional", "public service", "conformidade", "lei", "política", "politica", "norma", "institucional", "serviço público", "servico publico"],
    },
    "human_oversight_contestation": {
        "label": "Supervisão humana, reparo e contestação",
        "weight": 12,
        "strong": ["human oversight", "human-in-the-loop", "human in the loop", "human intervention", "contestability", "contestation", "redress", "appeal", "remedy", "handoff", "escalation", "supervisão humana", "supervisao humana", "intervenção humana", "intervencao humana", "contestabilidade", "contestação", "contestacao", "recurso", "reparação", "reparacao", "remediação", "remediacao", "transbordo", "escalonamento"],
        "weak": ["review", "challenge", "question", "explanation", "justification", "decision-maker", "revisão", "revisao", "questionar", "explicação", "explicacao", "justificativa", "tomador de decisão", "tomador de decisao"],
    },
    "risk_audit_compliance": {
        "label": "Risco, auditoria e compliance",
        "weight": 12,
        "strong": ["risk management", "risk assessment", "audit", "auditing", "audit trail", "compliance", "monitoring", "traceability", "logging", "logs", "controle de risco", "gestão de risco", "gestao de risco", "avaliação de risco", "avaliacao de risco", "auditoria", "trilha de auditoria", "conformidade", "monitoramento", "rastreabilidade", "registros"],
        "weak": ["risk", "safety", "security", "privacy", "harm", "failure", "mitigation", "risco", "segurança", "seguranca", "privacidade", "dano", "falha", "mitigação", "mitigacao"],
    },
    "operational_technical_governance": {
        "label": "Governança operacional e técnica",
        "weight": 8,
        "strong": ["rag", "retrieval augmented generation", "guardrail", "guardrails", "observability", "tracing", "post-deployment monitoring", "deployment monitoring", "knowledge base", "knowledge governance", "base de conhecimento", "governança do conhecimento", "governanca do conhecimento", "observabilidade", "rastreamento", "monitoramento pós-implantação", "monitoramento pos-implantacao"],
        "weak": ["deployment", "maintenance", "operation", "feedback loop", "continuous learning", "system performance", "implantação", "implantacao", "manutenção", "manutencao", "operação", "operacao", "ciclo de feedback", "aprendizagem contínua", "aprendizagem continua"],
    },
    "foundational_value": {
        "label": "Valor fundacional",
        "weight": 4,
        "strong": ["framework", "conceptual framework", "theoretical framework", "taxonomy", "model", "definition", "conceptual", "theory", "estrutura conceitual", "referencial teórico", "referencial teorico", "taxonomia", "modelo", "definição", "definicao", "conceitual", "teoria"],
        "weak": ["perspective", "lens", "map", "mapping", "principle", "perspectiva", "lente", "mapear", "mapeamento", "princípio", "principio"],
    },
}


@dataclass
class EvidenceHit:
    dimension: str
    label: str
    term: str
    term_strength: str
    page: int
    section: str
    excerpt: str
    directness_score: int
    relevance_score: int


# ============================================================
# UTILITÁRIOS
# ============================================================

def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def date_id() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", safe_text(text)).strip()


def normalize_string(value: str) -> str:
    import unicodedata
    value = safe_text(value).lower()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(c for c in value if not unicodedata.combining(c))
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return re.sub(r"_+", "_", value).strip("_")


def sha256_file(path: Path, block_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(block_size), b""):
            h.update(block)
    return h.hexdigest()


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8", errors="ignore")).hexdigest()


def short_hash(text: str, n: int = 12) -> str:
    return sha256_text(text)[:n]


def to_number(value: Any, default: float = 0.0) -> float:
    text = safe_text(value).replace(",", ".")
    if not text:
        return default
    try:
        return float(text)
    except Exception:
        return default


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def join_values(values: Iterable[Any], sep: str = " | ") -> str:
    out: List[str] = []
    for value in values:
        if isinstance(value, list):
            out.extend([safe_text(x) for x in value])
        else:
            out.append(safe_text(value))
    out = [x for x in out if x and x.lower() not in {"nan", "none", "null", "unclear"}]
    return sep.join(sorted(set(out)))


def grade_from_score(score: float) -> str:
    if score >= 85:
        return "A"
    if score >= 70:
        return "B"
    if score >= 55:
        return "C"
    if score >= 40:
        return "D"
    return "E"


# ============================================================
# PDF, SEÇÕES, METADADOS
# ============================================================

def extract_pdf_pages(pdf_path: Path) -> Tuple[List[Dict[str, Any]], int, str]:
    try:
        doc = fitz.open(str(pdf_path))
        pages: List[Dict[str, Any]] = []
        for i, page in enumerate(doc, start=1):
            text = page.get_text("text") or ""
            text = normalize_space(text)
            pages.append({"page": i, "text": text})
        page_count = len(doc)
        doc.close()
        total = sum(len(p["text"]) for p in pages)
        status = "success" if total >= 1000 else "partial" if total > 0 else "failed"
        return pages, page_count, status
    except Exception:
        return [], 0, "failed"


def pages_to_text(pages: List[Dict[str, Any]]) -> str:
    return "\n\n".join([f"[PAGE {p['page']}] {p['text']}" for p in pages if safe_text(p.get("text"))])


SECTION_PATTERNS: List[Tuple[str, str]] = [
    ("abstract", r"\babstract\b|\bresumo\b"),
    ("introduction", r"\bintroduction\b|\bintrodução\b|\bintroducao\b"),
    ("background", r"\bbackground\b|\bfundamentação\b|\bfundamentacao\b"),
    ("related_work", r"\brelated work\b|\btrabalhos relacionados\b"),
    ("methods", r"\bmethods\b|\bmethodology\b|\bmetodologia\b|\bmétodo\b|\bmetodo\b"),
    ("results", r"\bresults\b|\bresultados\b"),
    ("findings", r"\bfindings\b|\bachados\b"),
    ("discussion", r"\bdiscussion\b|\bdiscussão\b|\bdiscussao\b"),
    ("conclusion", r"\bconclusion\b|\bconclusions\b|\bconclusão\b|\bconclusao\b"),
    ("limitations", r"\blimitations\b|\blimitações\b|\blimitacoes\b"),
    ("references", r"\breferences\b|\bbibliography\b|\breferências\b|\breferencias\b"),
]


def infer_section_for_excerpt(text_before: str) -> str:
    window = text_before[-3500:].lower()
    last_section, last_pos = "unclear", -1
    for section, pattern in SECTION_PATTERNS:
        matches = list(re.finditer(pattern, window, flags=re.I))
        if matches and matches[-1].start() > last_pos:
            last_section, last_pos = section, matches[-1].start()
    return last_section


def references_start_index(text: str) -> Optional[int]:
    patterns = [
        r"\n\s*\[PAGE\s+\d+\]\s*(references|bibliography|referências|referencias)\b",
        r"\n\s*(references|bibliography|referências|referencias)\s*\n",
        r"\bReferences\b", r"\bBibliography\b", r"\bReferências\b", r"\bReferencias\b",
    ]
    candidates = []
    for pattern in patterns:
        for m in re.finditer(pattern, text, flags=re.I):
            if m.start() > len(text) * 0.35:
                candidates.append(m.start())
    return min(candidates) if candidates else None


def strip_references(text: str) -> Tuple[str, str]:
    idx = references_start_index(text)
    return (text[:idx], "yes") if idx is not None else (text, "unclear")


def first_n_pages_text(pages: List[Dict[str, Any]], n: int = 2) -> str:
    return "\n\n".join([p["text"] for p in pages[:n] if safe_text(p.get("text"))])


def possible_scanned_pdf(text: str, page_count: int) -> str:
    return "yes" if page_count <= 0 or (len(text) / max(page_count, 1)) < 300 else "no"


def detect_tables(text: str) -> str:
    return "yes" if re.search(r"\btable\s+\d+|\btabela\s+\d+", text, flags=re.I) else "unclear"


def detect_figures(text: str) -> str:
    return "yes" if re.search(r"\bfigure\s+\d+|\bfig\.\s*\d+|\bfigura\s+\d+", text, flags=re.I) else "unclear"


def extract_doi(text: str) -> str:
    matches = re.findall(r"\b10\.\d{4,9}/[-._;()/:A-Z0-9]+\b", text[:16000], flags=re.I)
    if not matches:
        matches = re.findall(r"\b10\.\d{4,9}/[-._;()/:A-Z0-9]+\b", text, flags=re.I)
    if not matches:
        return "unclear"
    return matches[0].rstrip(".,);]").replace("https://doi.org/", "").replace("http://dx.doi.org/", "")


def extract_year(front: str, full_text: str) -> str:
    candidates = re.findall(r"\b(19\d{2}|20\d{2})\b", front[:12000]) or re.findall(r"\b(19\d{2}|20\d{2})\b", full_text[:20000])
    years = [int(y) for y in candidates if 1900 <= int(y) <= datetime.now().year + 1]
    return str(Counter(years).most_common(1)[0][0]) if years else "unclear"


def extract_title(pdf_name: str, front: str) -> str:
    clean = normalize_space(re.sub(r"Downloaded from .*? Terms and Conditions.*?", " ", front, flags=re.I))
    for pattern in [
        r"(?:Research Article\s+\d+\s+)?(.{20,220}?)(?:\s+Abstract\s*:|\s+Abstract\b|\s+Resumo\s*:|\s+Resumo\b)",
        r"^(.{20,180}?)(?:\s+Authors?|\s+By\s+|\s+Abstract\s*:|\s+Resumo\s*:)",
    ]:
        m = re.search(pattern, clean, flags=re.I)
        if m:
            title = re.sub(r"^(article|research article|original article)\s+", "", normalize_space(m.group(1)), flags=re.I).strip(" .,:;")
            if 15 <= len(title) <= 220:
                return title
    return Path(pdf_name).stem


def extract_authors(front: str, title: str) -> str:
    clean = normalize_space(front)
    after_title = clean.split(title, 1)[-1] if title and title in clean else clean[:1500]
    before_abstract = re.split(r"\bAbstract\b|\bResumo\b", after_title, flags=re.I)[0][:700]
    candidate = normalize_space(re.sub(r"\S+@\S+", " ", re.sub(r"\bdoi\b.*", " ", before_abstract, flags=re.I)))
    bad = ["copyright", "creative commons", "open access", "downloaded from", "journal", "volume", "issue", "wiley online library", "research article"]
    if len(candidate) < 3 or any(b in candidate.lower() for b in bad):
        return "unclear"
    return candidate[:300] if re.search(r"\b[A-Z][a-z]+(?:\s+[A-Z]\.)?(?:\s+[A-Z][a-z]+)+\b", candidate) else "unclear"


def extract_venue(front: str) -> str:
    patterns = [r"Public Administration Review", r"Perspectives on Public Management and Governance", r"Journal of [A-Z][A-Za-z &]+", r"International Journal of [A-Z][A-Za-z &]+", r"Proceedings of [A-Z][A-Za-z0-9 &]+", r"ACM [A-Z][A-Za-z0-9 &]+", r"IEEE [A-Z][A-Za-z0-9 &]+"]
    for p in patterns:
        m = re.search(p, front, flags=re.I)
        if m:
            return normalize_space(m.group(0))
    return "unclear"


def infer_publication_type(front: str, doi: str, venue: str) -> str:
    lower = front.lower()
    if "preprint" in lower or "arxiv" in lower:
        return "preprint"
    if "conference" in lower or "proceedings" in lower or "acm" in lower or "ieee" in lower:
        return "conference_paper"
    if doi != "unclear" or "journal" in lower or venue != "unclear":
        return "journal_article"
    if "report" in lower:
        return "report"
    if "thesis" in lower or "dissertation" in lower:
        return "thesis"
    return "unknown"


def is_probably_english(text: str) -> bool:
    front = text[:6000].lower()
    return sum(1 for t in [" the ", " and ", " of ", " in ", " abstract", " introduction"] if t in front) >= 3


def is_probably_portuguese(text: str) -> bool:
    front = text[:6000].lower()
    return sum(1 for t in [" de ", " que ", " em ", " resumo", " introdução", " introducao"] if t in front) >= 3


def extract_bibliographic_metadata(pdf_name: str, pages: List[Dict[str, Any]], full_text: str) -> Dict[str, str]:
    front = first_n_pages_text(pages, 2)
    doi = extract_doi(front + "\n" + full_text[:12000])
    year = extract_year(front, full_text)
    title = extract_title(pdf_name, front)
    authors = extract_authors(front, title)
    venue = extract_venue(front)
    pub_type = infer_publication_type(front, doi, venue)
    language = "Portuguese" if is_probably_portuguese(front) else "English" if is_probably_english(front) else "unclear"
    return {"title": title, "authors": authors, "year": year, "venue": venue, "doi": doi, "publication_type": pub_type, "language": language}


# ============================================================
# TRIAGEM PYTHON E SNIPPETS
# ============================================================

def term_regex(term: str) -> re.Pattern:
    escaped = re.escape(term)
    if re.match(r"^[A-Za-z0-9_ \-]+$", term):
        pattern = r"(?<![A-Za-z0-9])" + escaped.replace(r"\ ", r"\s+") + r"(?![A-Za-z0-9])"
    else:
        pattern = escaped
    return re.compile(pattern, flags=re.I)


def directness_for_section(section: str) -> int:
    if section in {"abstract", "introduction", "methods", "results", "findings", "discussion", "conclusion"}:
        return 3
    if section in {"background", "related_work", "limitations"}:
        return 2
    return 0 if section == "references" else 1


def find_hits(dimension: str, label: str, terms: List[str], strength: str, pages: List[Dict[str, Any]], full_text: str) -> List[EvidenceHit]:
    hits: List[EvidenceHit] = []
    ref_pos = references_start_index(full_text)
    for page in pages:
        page_num = int(page["page"])
        page_text = safe_text(page.get("text"))
        page_token = f"[PAGE {page_num}] {page_text}"
        page_pos = full_text.find(page_token)
        if ref_pos is not None and page_pos >= ref_pos:
            continue
        for term in terms:
            for m in term_regex(term).finditer(page_text):
                start = max(0, m.start() - 260)
                end = min(len(page_text), m.end() + 260)
                section = infer_section_for_excerpt(page_text[:m.start()])
                if section == "references":
                    continue
                hits.append(EvidenceHit(dimension, label, term, strength, page_num, section, normalize_space(page_text[start:end])[:700], directness_for_section(section), 3 if strength == "strong" else 1))
                if len(hits) >= 12:
                    return hits
    return hits


def score_dimension(key: str, item: Dict[str, Any], pages: List[Dict[str, Any]], full_text: str) -> Tuple[float, int, List[EvidenceHit]]:
    strong = find_hits(key, item["label"], item.get("strong", []), "strong", pages, full_text)
    weak = find_hits(key, item["label"], item.get("weak", []), "weak", pages, full_text)
    hits = strong + weak
    sections = {h.section for h in hits}
    bonus = bool(sections.intersection({"abstract", "introduction", "methods", "results", "findings", "discussion", "conclusion"}))
    if len(strong) >= 3 or (len(strong) >= 2 and bonus):
        score_0_3 = 3
    elif len(strong) >= 1 or len(weak) >= 4:
        score_0_3 = 2
    elif len(weak) >= 1:
        score_0_3 = 1
    else:
        score_0_3 = 0
    return round((score_0_3 / 3) * float(item["weight"]), 4), score_0_3, hits


def python_triage(pages: List[Dict[str, Any]], full_text_no_refs: str) -> Dict[str, Any]:
    full_text = pages_to_text(pages)
    dim_scores: Dict[str, Any] = {}
    all_hits: List[EvidenceHit] = []
    total = 0.0
    for key, item in TAXONOMY.items():
        weighted, s03, hits = score_dimension(key, item, pages, full_text)
        dim_scores[key] = {"weighted": weighted, "score_0_3": s03, "matched_terms": sorted(set(h.term for h in hits)), "evidence_count": len(hits)}
        all_hits.extend(hits)
        total += weighted
    # traceability basic bonus
    total = round(clamp(total + 4.0, 0, 100), 2)
    grade = grade_from_score(total)
    llm = dim_scores["llm_genai"]["score_0_3"]
    conv = dim_scores["conversational_ai"]["score_0_3"]
    gov = dim_scores["governance_accountability"]["score_0_3"]
    reg = dim_scores["regulated_high_stakes"]["score_0_3"]
    risk = dim_scores["risk_audit_compliance"]["score_0_3"]
    foundational = dim_scores["foundational_value"]["score_0_3"]
    if total >= 70 and gov >= 2 and (llm >= 2 or conv >= 2) and (reg >= 1 or risk >= 1):
        decision = "central_candidate"
    elif total >= 55 and gov >= 2 and (llm >= 1 or conv >= 1 or reg >= 1 or risk >= 1):
        decision = "supporting_candidate"
    elif total >= 40 and gov >= 2 and foundational >= 2:
        decision = "foundational_contextual"
    elif total >= 30:
        decision = "borderline_manual"
    else:
        decision = "exclude_low_relevance"
    top = sorted(all_hits, key=lambda h: (h.relevance_score, h.directness_score, len(h.excerpt)), reverse=True)[0] if all_hits else None
    return {
        "total": total,
        "grade": grade,
        "decision": decision,
        "dimension_scores": dim_scores,
        "hits": all_hits,
        "top_evidence": top,
        "rationale": f"Score Python={total}/100; decisão={decision}; LLM={llm}/3; Conversacional={conv}/3; Governança={gov}/3; Regulado={reg}/3; Risco={risk}/3; Fundacional={foundational}/3.",
    }


def make_llm_context(pages: List[Dict[str, Any]], body_no_refs: str, py_result: Dict[str, Any], max_chars: int) -> str:
    full = pages_to_text(pages)
    # Prioridade: front matter, abstract/início, melhores evidências, métodos/resultados/discussão/conclusão.
    front = full[: min(len(full), int(max_chars * 0.30))]
    evidence = "\n\n".join([f"[PAGE {h.page} | {h.section} | term={h.term}] {h.excerpt}" for h in py_result["hits"][:30]])
    # Pega blocos intermediários e final sem referências.
    middle_start = max(0, len(body_no_refs) // 2 - int(max_chars * 0.12))
    middle = body_no_refs[middle_start: middle_start + int(max_chars * 0.24)]
    tail = body_no_refs[-int(max_chars * 0.22):]
    ctx = f"[FRONT MATTER]\n{front}\n\n[PYTHON EVIDENCE HITS]\n{evidence}\n\n[MIDDLE]\n{middle}\n\n[END BEFORE REFERENCES]\n{tail}"
    return ctx[:max_chars]


# ============================================================
# LLM CLIENT
# ============================================================

class LLMClient:
    def __init__(
        self,
        provider: str,
        api_base: str,
        api_key: str,
        model: str,
        timeout: int = 120,
        temperature: float = 0.0,
        retries: int = 2,
        cursor_command: str = "agent",
        cursor_pass_mode: str = "stdin",
    ):
        self.provider = provider
        self.api_base = api_base.rstrip("/")
        self.api_key = api_key
        self.model = model
        self.timeout = timeout
        self.temperature = temperature
        self.retries = retries
        self.cursor_command = cursor_command
        self.cursor_pass_mode = cursor_pass_mode

    def chat_json(self, system_prompt: str, user_prompt: str) -> Dict[str, Any]:
        last_error: Optional[Exception] = None
        for attempt in range(self.retries + 1):
            try:
                print(f"      tentativa LLM {attempt+1}/{self.retries+1}...", flush=True)
                if self.provider == "ollama":
                    return self._chat_ollama(system_prompt, user_prompt)
                if self.provider == "cursor-cli":
                    return self._chat_cursor_cli(system_prompt, user_prompt)
                return self._chat_openai_compatible(system_prompt, user_prompt)
            except Exception as exc:
                last_error = exc
                if attempt < self.retries:
                    time.sleep(2 + attempt * 4)
                else:
                    raise RuntimeError(f"LLM falhou após {self.retries + 1} tentativas: {last_error}")
        raise RuntimeError(f"LLM falhou: {last_error}")

    def _chat_openai_compatible(self, system_prompt: str, user_prompt: str) -> Dict[str, Any]:
        url = f"{self.api_base}/chat/completions"
        body = {
            "model": self.model,
            "messages": [{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],
            "temperature": self.temperature,
            "response_format": {"type": "json_object"},
        }
        data = json.dumps(body).encode("utf-8")
        headers = {"Content-Type": "application/json", "User-Agent": "metagrade-python-llm-adjudicator/1.1"}
        if "openrouter.ai" in self.api_base:
            headers["HTTP-Referer"] = "http://localhost"
            headers["X-Title"] = "MetaGrade Python LLM Adjudicator"
        if self.api_key:
            headers["Authorization"] = f"Bearer {self.api_key}"
        req = urllib.request.Request(url, data=data, headers=headers, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=self.timeout) as resp:
                raw = resp.read().decode("utf-8", errors="ignore")
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="ignore")
            raise RuntimeError(f"HTTP {exc.code}: {detail[:1000]}")
        payload = json.loads(raw)
        content = payload.get("choices", [{}])[0].get("message", {}).get("content", "")
        return parse_json_object(content)

    def _chat_ollama(self, system_prompt: str, user_prompt: str) -> Dict[str, Any]:
        url = f"{self.api_base}/api/chat"
        body = {
            "model": self.model,
            "messages": [{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],
            "stream": False,
            "format": "json",
            "options": {"temperature": self.temperature, "num_ctx": 32768},
        }
        data = json.dumps(body).encode("utf-8")
        req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"}, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=self.timeout) as resp:
                raw = resp.read().decode("utf-8", errors="ignore")
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="ignore")
            raise RuntimeError(f"HTTP {exc.code}: {detail[:1000]}")
        payload = json.loads(raw)
        content = payload.get("message", {}).get("content", "")
        return parse_json_object(content)

    def _chat_cursor_cli(self, system_prompt: str, user_prompt: str) -> Dict[str, Any]:
        """
        Chama o Cursor CLI em modo headless para usar Composer 2.5.

        Requisitos externos:
        - Cursor CLI instalado e autenticado;
        - comando disponível como `agent` ou `cursor-agent`;
        - modelo aceito pelo Cursor CLI, por exemplo `composer-2.5`.

        Observação: isto não usa OpenAI API. O Python apenas invoca o processo do Cursor CLI.
        """
        command = resolve_cursor_cli_command(self.cursor_command)

        combined_prompt = (
            "Você está executando uma avaliação científica estruturada. "
            "Responda exclusivamente com JSON válido, sem markdown, sem comentários e sem texto fora do JSON.\n\n"
            "[SYSTEM]\n" + system_prompt + "\n\n[USER]\n" + user_prompt
        )

        # --trust evita o prompt interativo de "Workspace Trust", que travaria
        # a execução headless deste pipeline.
        base_cmd = [command, "-p", "--model", self.model, "--trust"]
        if self.api_key:
            base_cmd.extend(["--api-key", self.api_key])

        attempts: List[Tuple[str, List[str], Optional[str]]] = []
        if self.cursor_pass_mode == "arg":
            # Pode bater limite de tamanho no Windows se max_chars for alto.
            attempts.append(("arg", base_cmd + [combined_prompt], None))
            attempts.append(("stdin", base_cmd, combined_prompt))
        else:
            attempts.append(("stdin", base_cmd, combined_prompt))
            if len(combined_prompt) < 7000:
                attempts.append(("arg", base_cmd + [combined_prompt], None))

        last_error = ""
        for mode, cmd, stdin_text in attempts:
            try:
                print(f"      Cursor CLI: {' '.join(cmd[:4])} ... modo={mode}", flush=True)
                proc = subprocess.run(
                    cmd,
                    input=stdin_text,
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="ignore",
                    timeout=self.timeout,
                    # No Windows o Cursor CLI é um shim .cmd/.ps1, não um .exe;
                    # subprocess precisa do shell para conseguir executá-lo.
                    shell=(os.name == "nt"),
                )
                stdout = proc.stdout or ""
                stderr = proc.stderr or ""
                if proc.returncode != 0:
                    last_error = f"Cursor CLI retornou código {proc.returncode}. STDERR: {stderr[-1500:]} STDOUT: {stdout[-1500:]}"
                    continue
                if not stdout.strip():
                    last_error = f"Cursor CLI não retornou conteúdo. STDERR: {stderr[-1500:]}"
                    continue
                return parse_json_object(stdout)
            except subprocess.TimeoutExpired:
                last_error = f"Cursor CLI excedeu timeout de {self.timeout}s. Reduza --max-chars ou aumente --timeout."
            except Exception as exc:
                last_error = f"Erro chamando Cursor CLI: {exc}"

        raise RuntimeError(last_error or "Cursor CLI falhou sem detalhe.")


def parse_json_object(text: str) -> Dict[str, Any]:
    text = safe_text(text)
    text = re.sub(r"^```(?:json)?", "", text, flags=re.I).strip()
    text = re.sub(r"```$", "", text).strip()
    text = re.sub(r"<think>.*?</think>", "", text, flags=re.S | re.I).strip()
    try:
        return json.loads(text)
    except Exception:
        start, end = text.find("{"), text.rfind("}")
        if start >= 0 and end >= start:
            return json.loads(text[start:end + 1])
        raise


def build_system_prompt(profile: str) -> str:
    return f"""
Você é um avaliador de revisão sistemática de literatura, operando como apoio de adjudicação assistida por LLM.
Perfil desta rodada: {profile}.

Escopo do projeto:
{PROJECT_SCOPE}

Regras obrigatórias:
1. Use somente o texto e as evidências fornecidas. Não invente autores, ano, DOI, método, achados ou contexto.
2. Não use o escopo do projeto como evidência do artigo. Evidência válida vem do texto do PDF.
3. Se algo estiver ausente, use "unclear".
4. Diferencie estudo central, estudo de apoio, referência fundacional/contextual e exclusão.
5. Para appraisal CASP/JBI adaptado, avalie apenas sinais presentes no texto fornecido.
6. Em cada evidência, use trecho literal curto e informe a página quando disponível.
7. Responda somente JSON válido, sem markdown.
""".strip()


def build_user_prompt(meta: Dict[str, str], py_result: Dict[str, Any], context: str) -> str:
    schema_hint = {
        "bibliographic": {"title": "", "authors": "", "year": "", "venue": "", "doi": "", "publication_type": "journal_article|conference_paper|preprint|report|thesis|book|unknown", "peer_reviewed": "yes|no|unclear", "citation_apa": ""},
        "methodology": {"research_type": "empirical|conceptual|theoretical|methodological|technical|review|mixed|unknown", "method_family": "", "study_design": "", "sample_description": "", "sample_size": "", "sector": "", "regulated_domain": "yes|no|partial|unclear", "llm_specificity": "specific_llm|generic_genai|pre_llm_ai|not_applicable|unclear", "technology_context": ""},
        "eligibility": {"title_abstract_decision": "include|exclude|maybe|unclear", "full_text_decision": "include|exclude|maybe|unclear", "final_decision": "central_evidence|supporting_evidence|foundational_contextual|borderline|exclude", "prisma_stage": "", "inclusion_codes": [], "exclusion_codes": [], "exclusion_reason": "", "decision_rationale": "", "decision_confidence": 0.0, "manual_spotcheck_required": "yes|no", "manual_spotcheck_reason": ""},
        "appraisal": {"objective_clarity": 0, "method_fit": 0, "sampling_adequacy": 0, "data_collection_transparency": 0, "analysis_rigor": 0, "ethics_reflexivity": 0, "results_clarity": 0, "limitations_transparency": 0, "evidence_support": 0, "transferability": 0, "notes": ""},
        "cerqual": {"methodological_limitations": 0, "coherence": 0, "adequacy": 0, "relevance": 0, "overall_confidence": "high|moderate|low|very_low|unclear", "explanation": ""},
        "coding": {"open_codes": [], "axial_codes": [], "theme": "", "subtheme": "", "concepts": [], "model_layers": [], "rq_alignment": [], "coding_confidence": 0.0},
        "synthesis": {"main_claim": "", "contribution": "", "limitations": "", "use_in_review": ""},
        "evidence": [{"claim": "", "excerpt": "", "page": "", "section": "", "supports": "eligibility|appraisal|coding|synthesis"}],
    }
    return f"""
Metadados extraídos por Python, ainda sujeitos a correção:
{json.dumps(meta, ensure_ascii=False)}

Resultado da triagem determinística Python:
{json.dumps({k:v for k,v in py_result.items() if k not in ['hits', 'top_evidence']}, ensure_ascii=False, default=str)}

Schema esperado:
{json.dumps(schema_hint, ensure_ascii=False)}

Texto e evidências do PDF:
{context}
""".strip()


# ============================================================
# VALIDAÇÃO DE EVIDÊNCIAS E CONSENSO
# ============================================================

def evidence_exists(excerpt: str, full_text: str) -> bool:
    ex = normalize_space(excerpt)
    if not ex or ex.lower() == "unclear":
        return False
    full = normalize_space(full_text)
    if ex[:160] in full:
        return True
    # Tolerância: checa blocos de 80 chars quando a LLM encurta com reticências.
    chunks = [c.strip() for c in re.split(r"\.\.\.|…", ex) if len(c.strip()) >= 60]
    return any(c[:80] in full for c in chunks)


def validate_llm_evidence(llm_data: Dict[str, Any], full_text: str) -> Tuple[int, int, str]:
    items = llm_data.get("evidence", []) or []
    valid, invalid = 0, 0
    notes = []
    for idx, item in enumerate(items):
        ex = safe_text(item.get("excerpt")) if isinstance(item, dict) else ""
        if evidence_exists(ex, full_text):
            valid += 1
        else:
            invalid += 1
            notes.append(f"evidence[{idx}] not found: {ex[:120]}")
    return valid, invalid, " | ".join(notes[:10])


def appraisal_total(appraisal: Dict[str, Any]) -> float:
    keys = ["objective_clarity", "method_fit", "sampling_adequacy", "data_collection_transparency", "analysis_rigor", "ethics_reflexivity", "results_clarity", "limitations_transparency", "evidence_support", "transferability"]
    return round(sum(clamp(to_number(appraisal.get(k)), 0, 3) for k in keys), 2)


def appraisal_grade(total: float) -> str:
    if total >= 24:
        return "high"
    if total >= 18:
        return "moderate"
    if total >= 12:
        return "limited"
    return "low"


def consensus_score(py_total: float, llm_decision: str, confidence: float, valid_evidence: int, invalid_evidence: int, appraisal_sum: float) -> float:
    base = py_total * 0.45
    decision_bonus = {
        "central_evidence": 25,
        "supporting_evidence": 18,
        "foundational_contextual": 13,
        "borderline": 8,
        "exclude": 0,
    }.get(llm_decision, 5)
    evidence_bonus = min(valid_evidence, 5) * 2
    appraisal_bonus = (appraisal_sum / 30) * 15
    confidence_bonus = clamp(confidence, 0, 1) * 10
    penalty = min(invalid_evidence * 3, 15)
    return round(clamp(base + decision_bonus + evidence_bonus + appraisal_bonus + confidence_bonus - penalty, 0, 100), 2)


def consensus_priority(decision: str, score: float) -> int:
    if decision == "central_evidence" or score >= 80:
        return 1
    if decision == "supporting_evidence" or score >= 65:
        return 2
    if decision == "foundational_contextual" or score >= 50:
        return 3
    if decision == "borderline" or score >= 35:
        return 4
    return 5


# ============================================================
# PROCESSAMENTO DE PDF
# ============================================================

def process_pdf(pdf_path: Path, output_dir: Path, run_id: str, client: LLMClient, args: argparse.Namespace) -> Dict[str, Any]:
    raw_dir = output_dir / "llm_raw_outputs"
    text_dir = output_dir / "extracted_text"
    raw_dir.mkdir(parents=True, exist_ok=True)
    text_dir.mkdir(parents=True, exist_ok=True)

    pdf_hash = sha256_file(pdf_path)
    pages, page_count, extraction_status = extract_pdf_pages(pdf_path)
    full_text = pages_to_text(pages)
    text_hash = sha256_text(full_text)
    body_no_refs, refs_detected = strip_references(full_text)
    meta = extract_bibliographic_metadata(pdf_path.name, pages, full_text)
    py_result = python_triage(pages, body_no_refs)

    text_path = text_dir / f"{pdf_path.stem}_{pdf_hash[:10]}.txt"
    if args.write_text:
        text_path.write_text(full_text, encoding="utf-8")

    context = make_llm_context(pages, body_no_refs, py_result, args.max_chars)
    profiles = ["methodological_reviewer", "scope_reviewer", "critical_evidence_reviewer"]

    llm_outputs: List[Dict[str, Any]] = []
    for r in range(args.reviewers):
        raw_path = raw_dir / f"{pdf_path.stem}_{pdf_hash[:10]}_reviewer{r+1}.json"
        if args.resume and raw_path.exists():
            data = json.loads(raw_path.read_text(encoding="utf-8"))
        else:
            print(f"    LLM reviewer {r+1}/{args.reviewers}: chamando {client.provider}/{client.model}...", flush=True)
            data = client.chat_json(build_system_prompt(profiles[r % len(profiles)]), build_user_prompt(meta, py_result, context))
            print(f"    LLM reviewer {r+1}/{args.reviewers}: OK", flush=True)
            raw_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        data["_reviewer_profile"] = profiles[r % len(profiles)]
        llm_outputs.append(data)

    # Consolidação simples: usa primeiro revisor como base e reforça com médias/maior confiança quando múltiplos.
    # Para artigo completo, normalmente 1 revisor LLM + validação de evidência é mais custo-efetivo.
    data = llm_outputs[0]
    if len(llm_outputs) > 1:
        # escolhe output com maior decision_confidence
        data = sorted(llm_outputs, key=lambda d: to_number(((d.get("eligibility") or {}).get("decision_confidence"))), reverse=True)[0]

    valid_e, invalid_e, invalid_notes = validate_llm_evidence(data, full_text)

    b = data.get("bibliographic", {}) or {}
    m = data.get("methodology", {}) or {}
    e = data.get("eligibility", {}) or {}
    app = data.get("appraisal", {}) or {}
    cerq = data.get("cerqual", {}) or {}
    cod = data.get("coding", {}) or {}
    syn = data.get("synthesis", {}) or {}

    app_total = appraisal_total(app)
    final_decision = safe_text(e.get("final_decision")) or "borderline"
    conf = clamp(to_number(e.get("decision_confidence")), 0, 1)
    c_score = consensus_score(py_result["total"], final_decision, conf, valid_e, invalid_e, app_total)
    c_grade = grade_from_score(c_score)
    c_priority = consensus_priority(final_decision, c_score)

    study_id_base = safe_text(b.get("doi")) if safe_text(b.get("doi")) and safe_text(b.get("doi")).lower() != "unclear" else f"{b.get('title') or meta.get('title')}|{b.get('year') or meta.get('year')}|{pdf_hash[:12]}"
    study_id = "S_" + short_hash(study_id_base, 12)
    record_id = "R_" + short_hash(f"{study_id}|{pdf_hash}|{run_id}", 14)

    top = py_result.get("top_evidence")
    top_excerpt = top.excerpt if top else "unclear"
    top_page = str(top.page) if top else "unclear"
    top_section = top.section if top else "unclear"

    flags: List[str] = []
    if extraction_status != "success": flags.append("text_extraction_not_success")
    if possible_scanned_pdf(full_text, page_count) == "yes": flags.append("possible_scanned_pdf")
    if invalid_e > 0: flags.append("llm_evidence_not_found")
    if valid_e == 0: flags.append("no_valid_llm_evidence")
    if conf < 0.55: flags.append("low_llm_confidence")
    if final_decision in {"central_evidence", "supporting_evidence"} and invalid_e > valid_e: flags.append("decision_evidence_conflict")
    if len(llm_outputs) > 1:
        decisions = [safe_text((d.get("eligibility") or {}).get("final_decision")) for d in llm_outputs]
        if len(set(decisions)) > 1:
            flags.append("reviewer_disagreement")

    raw_output_path = str(raw_dir / f"{pdf_path.stem}_{pdf_hash[:10]}_reviewer1.json")

    row = {col: "" for col in OUTPUT_COLUMNS}
    row.update({
        "record_id": record_id, "study_id": study_id, "run_id": run_id, "script_version": SCRIPT_VERSION, "created_at": now_str(),
        "raw_file_name": pdf_path.name, "pdf_path": str(pdf_path), "pdf_hash": pdf_hash, "text_hash": text_hash, "page_count": page_count,
        "text_extraction_status": extraction_status, "text_char_count": len(full_text), "body_char_count_no_references": len(body_no_refs),
        "possible_scanned_pdf": possible_scanned_pdf(full_text, page_count), "references_detected": refs_detected, "tables_detected": detect_tables(full_text), "figures_detected": detect_figures(full_text),
        "extracted_text_path": str(text_path) if args.write_text else "not_written", "llm_raw_output_path": raw_output_path,
        "py_title": meta.get("title"), "py_authors": meta.get("authors", "unclear"), "py_year": meta.get("year"), "py_venue": meta.get("venue"), "py_doi": meta.get("doi"), "py_publication_type": meta.get("publication_type"), "py_language": meta.get("language"),
        "py_scope_total_score": py_result["total"], "py_scope_grade": py_result["grade"], "py_triage_decision": py_result["decision"], "py_triage_rationale": py_result["rationale"],
        "py_matched_terms_summary": json.dumps(py_result["dimension_scores"], ensure_ascii=False), "py_top_evidence_excerpt": top_excerpt, "py_top_evidence_page": top_page, "py_top_evidence_section": top_section,

        "llm_title": b.get("title", "unclear"), "llm_authors": b.get("authors", "unclear"), "llm_year": b.get("year", "unclear"), "llm_venue": b.get("venue", "unclear"), "llm_doi": b.get("doi", "unclear"), "llm_publication_type": b.get("publication_type", "unclear"), "llm_peer_reviewed": b.get("peer_reviewed", "unclear"), "llm_citation_apa": b.get("citation_apa", "unclear"),
        "llm_research_type": m.get("research_type", "unclear"), "llm_method_family": m.get("method_family", "unclear"), "llm_study_design": m.get("study_design", "unclear"), "llm_sample_description": m.get("sample_description", "unclear"), "llm_sample_size": m.get("sample_size", "unclear"), "llm_sector": m.get("sector", "unclear"), "llm_regulated_domain": m.get("regulated_domain", "unclear"), "llm_llm_specificity": m.get("llm_specificity", "unclear"), "llm_technology_context": m.get("technology_context", "unclear"),
        "llm_title_abstract_decision": e.get("title_abstract_decision", "unclear"), "llm_full_text_decision": e.get("full_text_decision", "unclear"), "llm_final_decision": final_decision, "llm_prisma_stage": e.get("prisma_stage", "full_text_llm_assessed"), "llm_inclusion_codes": join_values(e.get("inclusion_codes", [])), "llm_exclusion_codes": join_values(e.get("exclusion_codes", [])), "llm_exclusion_reason": e.get("exclusion_reason", ""), "llm_decision_rationale": e.get("decision_rationale", ""), "llm_decision_confidence": conf, "llm_manual_spotcheck_required": e.get("manual_spotcheck_required", "yes"), "llm_manual_spotcheck_reason": e.get("manual_spotcheck_reason", ""),
        "appraisal_objective_clarity": app.get("objective_clarity", 0), "appraisal_method_fit": app.get("method_fit", 0), "appraisal_sampling_adequacy": app.get("sampling_adequacy", 0), "appraisal_data_collection_transparency": app.get("data_collection_transparency", 0), "appraisal_analysis_rigor": app.get("analysis_rigor", 0), "appraisal_ethics_reflexivity": app.get("ethics_reflexivity", 0), "appraisal_results_clarity": app.get("results_clarity", 0), "appraisal_limitations_transparency": app.get("limitations_transparency", 0), "appraisal_evidence_support": app.get("evidence_support", 0), "appraisal_transferability": app.get("transferability", 0), "appraisal_total_score": app_total, "appraisal_quality_grade": appraisal_grade(app_total), "appraisal_notes": app.get("notes", ""),
        "cerqual_methodological_limitations": cerq.get("methodological_limitations", 0), "cerqual_coherence": cerq.get("coherence", 0), "cerqual_adequacy": cerq.get("adequacy", 0), "cerqual_relevance": cerq.get("relevance", 0), "cerqual_overall_confidence": cerq.get("overall_confidence", "unclear"), "cerqual_explanation": cerq.get("explanation", ""),
        "coding_open_codes": join_values(cod.get("open_codes", [])), "coding_axial_codes": join_values(cod.get("axial_codes", [])), "coding_theme": cod.get("theme", ""), "coding_subtheme": cod.get("subtheme", ""), "coding_concepts": join_values(cod.get("concepts", [])), "coding_model_layers": join_values(cod.get("model_layers", [])), "coding_rq_alignment": join_values(cod.get("rq_alignment", [])), "coding_confidence": cod.get("coding_confidence", 0),
        "synthesis_main_claim": syn.get("main_claim", ""), "synthesis_contribution": syn.get("contribution", ""), "synthesis_limitations": syn.get("limitations", ""), "synthesis_use_in_review": syn.get("use_in_review", ""),
        "llm_evidence_json": json.dumps(data.get("evidence", []), ensure_ascii=False), "validated_evidence_count": valid_e, "invalid_evidence_count": invalid_e, "invalid_evidence_notes": invalid_notes,
        "consensus_final_decision": final_decision, "consensus_score": c_score, "consensus_grade": c_grade, "consensus_priority": c_priority, "consensus_rationale": f"Combina score Python, decisão LLM, confiança, evidências válidas e appraisal. Python={py_result['total']}; LLM={final_decision}; conf={conf}; valid_evidence={valid_e}; invalid_evidence={invalid_e}; appraisal={app_total}.",
        "data_quality_flags": "|".join(sorted(set(flags))), "audit_status": "needs_attention" if flags else "checked", "audit_notes": " | ".join(sorted(set(flags))) if flags else "no critical flags", "notes_for_reproducibility": "Python deterministic triage + LLM-assisted adjudication with literal evidence validation.",
    })
    return row


# ============================================================
# CHECKPOINT, XLSX E SANITIZAÇÃO
# ============================================================

EXCEL_MAX_CELL_CHARS = 32767


def sanitize_xml_text(value: Any, max_chars: int = EXCEL_MAX_CELL_CHARS) -> str:
    import unicodedata
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple, set)):
        text = json.dumps(value, ensure_ascii=False)
    else:
        text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    chars = []
    for ch in text:
        code = ord(ch)
        if ch in "\t\n\r":
            chars.append(ch); continue
        valid = (0x20 <= code <= 0xD7FF) or (0xE000 <= code <= 0xFFFD) or (0x10000 <= code <= 0x10FFFF)
        if not valid or unicodedata.category(ch) in {"Cc", "Cf", "Cs"}:
            chars.append(" ")
        else:
            chars.append(ch)
    out = "".join(chars).replace("\x00", " ")
    out = re.sub(r"[ \t]{2,}", " ", out)
    out = re.sub(r"\n{4,}", "\n\n\n", out)
    if len(out) > max_chars:
        out = out[:max_chars - 80] + " ... [TRUNCADO PARA LIMITE DO EXCEL]"
    return out


def sanitize_excel_value(value: Any) -> Any:
    if value is None:
        return ""
    try:
        if pd.isna(value): return ""
    except Exception:
        pass
    if isinstance(value, bool): return value
    if isinstance(value, int) and not isinstance(value, bool): return value
    if isinstance(value, float):
        return "" if math.isnan(value) or math.isinf(value) else value
    return sanitize_xml_text(value)


def sanitize_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    clean = df.copy()
    clean.columns = [sanitize_xml_text(c, max_chars=255) for c in clean.columns]
    for col in clean.columns:
        clean[col] = clean[col].map(sanitize_excel_value)
    return clean


def sanitize_sheet_name(name: str) -> str:
    name = re.sub(r"[\[\]\:\*\?\/\\]", "_", sanitize_xml_text(name, max_chars=31))
    return (name.strip() or "sheet")[:31]


def load_checkpoint(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    rows = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            if line.strip():
                rows.append(json.loads(line))
    return rows


def append_checkpoint(path: Path, row: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(row, ensure_ascii=False) + "\n")


def mark_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    seen: Dict[str, str] = {}
    keys, isdup, dupof = [], [], []
    for _, row in df.iterrows():
        doi = safe_text(row.get("llm_doi") or row.get("py_doi")).lower()
        title = normalize_string(safe_text(row.get("llm_title") or row.get("py_title")))
        year = safe_text(row.get("llm_year") or row.get("py_year"))
        key = f"doi:{doi}" if doi and doi != "unclear" else f"title_year:{title}:{year}" if title else f"file:{safe_text(row.get('pdf_hash'))}"
        keys.append(key)
        if key in seen:
            isdup.append("yes"); dupof.append(seen[key])
        else:
            isdup.append("no"); dupof.append(""); seen[key] = safe_text(row.get("record_id"))
    df["dedup_key"] = keys; df["is_duplicate"] = isdup; df["duplicate_of"] = dupof
    return df


def make_tabs(df: pd.DataFrame, manifest: Dict[str, Any]) -> Dict[str, pd.DataFrame]:
    df = mark_duplicates(df)
    studies_cols = ["study_id", "llm_title", "llm_authors", "llm_year", "llm_venue", "llm_doi", "llm_research_type", "llm_method_family", "llm_final_decision", "consensus_score", "consensus_grade", "consensus_priority", "cerqual_overall_confidence", "data_quality_flags", "pdf_path"]
    queue_cols = ["consensus_priority", "consensus_score", "llm_final_decision", "llm_title", "llm_authors", "llm_year", "llm_decision_rationale", "synthesis_use_in_review", "invalid_evidence_notes", "data_quality_flags", "pdf_path"]
    appraisal_cols = ["study_id", "llm_title", "appraisal_total_score", "appraisal_quality_grade", "appraisal_objective_clarity", "appraisal_method_fit", "appraisal_sampling_adequacy", "appraisal_data_collection_transparency", "appraisal_analysis_rigor", "appraisal_ethics_reflexivity", "appraisal_results_clarity", "appraisal_limitations_transparency", "appraisal_evidence_support", "appraisal_transferability", "appraisal_notes"]
    coding_cols = ["study_id", "llm_title", "coding_theme", "coding_subtheme", "coding_open_codes", "coding_axial_codes", "coding_concepts", "coding_model_layers", "coding_rq_alignment", "coding_confidence"]

    evidence_rows = []
    for _, row in df.iterrows():
        raw = safe_text(row.get("llm_evidence_json"))
        try:
            items = json.loads(raw) if raw else []
        except Exception:
            items = []
        for item in items:
            if isinstance(item, dict):
                evidence_rows.append({"study_id": row.get("study_id"), "title": row.get("llm_title"), "decision": row.get("llm_final_decision"), **item})

    prisma_rows = []
    for col in ["llm_title_abstract_decision", "llm_full_text_decision", "llm_final_decision", "consensus_final_decision", "audit_status", "is_duplicate"]:
        if col in df.columns:
            for value, count in df[col].fillna("blank").replace("", "blank").value_counts().items():
                prisma_rows.append({"dimension": col, "value": value, "count": int(count)})

    manifest_df = pd.DataFrame([{"key": k, "value": json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v} for k, v in manifest.items()])
    return {
        "records_flat": df,
        "studies": df[[c for c in studies_cols if c in df.columns]].sort_values(["consensus_priority", "consensus_score"], ascending=[True, False]),
        "evidence_matrix": pd.DataFrame(evidence_rows),
        "appraisal_casp_jbi": df[[c for c in appraisal_cols if c in df.columns]],
        "coding_themes": df[[c for c in coding_cols if c in df.columns]],
        "manual_spotcheck_queue": df[(df.get("audit_status", "") == "needs_attention") | (df.get("llm_manual_spotcheck_required", "") == "yes")][[c for c in queue_cols if c in df.columns]].sort_values(["consensus_priority", "consensus_score"], ascending=[True, False]),
        "prisma_summary": pd.DataFrame(prisma_rows),
        "run_manifest": manifest_df,
    }


def autosize(workbook_path: Path) -> None:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(workbook_path)
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            if ws.max_row >= 1 and ws.max_column >= 1:
                ws.auto_filter.ref = ws.dimensions
            for col in ws.columns:
                letter = col[0].column_letter
                width = min(max(max((len(str(c.value)) if c.value is not None else 0) for c in col) + 2, 10), 80)
                ws.column_dimensions[letter].width = width
        wb.save(workbook_path)
    except Exception as exc:
        print(f"Aviso: não consegui formatar XLSX: {exc}")


def write_xlsx(rows: List[Dict[str, Any]], output_dir: Path, manifest: Dict[str, Any]) -> None:
    df = pd.DataFrame(rows)
    for c in OUTPUT_COLUMNS:
        if c not in df.columns: df[c] = ""
    df = df[OUTPUT_COLUMNS + [c for c in df.columns if c not in OUTPUT_COLUMNS]]
    tabs = make_tabs(df, manifest)
    path = output_dir / "metagrade_python_llm_workbook.xlsx"
    if path.exists():
        try: path.unlink()
        except PermissionError: path = output_dir / f"metagrade_python_llm_workbook_{manifest['run_id']}.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, tab in tabs.items():
            sanitize_dataframe_for_excel(tab).to_excel(writer, sheet_name=sanitize_sheet_name(name), index=False)
    autosize(path)
    print(f"\nArquivo gerado: {path}")


# ============================================================
# MAIN
# ============================================================


def resolve_cursor_cli_command(preferred: str = "agent") -> str:
    """
    Resolve o comando real do Cursor CLI antes de processar PDFs.

    A página oficial do Cursor CLI mostra o comando `agent` como entrada principal.
    Algumas instalações antigas ou wrappers podem expor `cursor-agent`; por isso
    mantemos fallback. Se nenhum existir no PATH, falha rápido para evitar gerar
    408 linhas de erro.
    """
    candidates: List[str] = []
    if preferred:
        candidates.append(preferred)
    for c in ["agent", "cursor-agent"]:
        if c not in candidates:
            candidates.append(c)
    for c in candidates:
        found = shutil.which(c)
        if found:
            return c
    raise SystemExit(
        "Cursor CLI não encontrado no PATH. O script foi interrompido antes de processar PDFs para evitar um XLSX cheio de erros.\n\n"
        "Valide no PowerShell:\n"
        "  where agent\n"
        "  agent --version\n\n"
        "Se o comando não existir, instale/ative o Cursor CLI no Cursor. A documentação oficial mostra o instalador: \n"
        "  curl https://cursor.com/install -fsS | bash\n\n"
        "No Windows, instale pelo fluxo do próprio Cursor/CLI e reabra o PowerShell para atualizar o PATH. "
        "Depois rode novamente com --preflight."
    )


def cursor_cli_preflight_or_exit(args: argparse.Namespace) -> None:
    """Falha rápido quando provider=cursor-cli e o comando não existe."""
    if args.provider != "cursor-cli":
        return
    resolved = resolve_cursor_cli_command(args.cursor_command)
    if resolved != args.cursor_command:
        print(f"Cursor CLI encontrado como '{resolved}' (substituindo --cursor-command {args.cursor_command!r}).", flush=True)
        args.cursor_command = resolved
    else:
        print(f"Cursor CLI encontrado: {resolved}", flush=True)


def collect_pdfs(input_path: Path, limit: int = 0) -> List[Path]:
    if input_path.is_file() and input_path.suffix.lower() == ".pdf":
        pdfs = [input_path]
    elif input_path.is_dir():
        pdfs = sorted(input_path.glob("*.pdf"))
    else:
        raise SystemExit(f"Input inválido: {input_path}")
    return pdfs[:limit] if limit and limit > 0 else pdfs


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="MetaGrade Python + LLM Adjudicator")
    p.add_argument("--input", default=DEFAULT_INPUT_DIR)
    p.add_argument("--output", default=DEFAULT_OUTPUT_DIR)
    p.add_argument("--provider", choices=["cursor-cli", "openai", "ollama"], default="cursor-cli")
    p.add_argument("--api-base", default=os.getenv("LLM_API_BASE", "https://api.openai.com/v1"))
    p.add_argument("--api-key", default=os.getenv("LLM_API_KEY", os.getenv("OPENAI_API_KEY", "")))
    p.add_argument("--model", default=os.getenv("LLM_MODEL", "composer-2.5"))
    p.add_argument("--timeout", type=int, default=180, help="Timeout por chamada LLM/CLI em segundos.")
    p.add_argument("--temperature", type=float, default=0.0)
    p.add_argument("--retries", type=int, default=1, help="Retentativas por chamada LLM. Use 0 para falhar rápido.")
    p.add_argument("--reviewers", type=int, default=1, help="1 é recomendado para custo. Use 2 ou 3 para maior robustez.")
    p.add_argument("--max-chars", type=int, default=12000, help="Máximo de caracteres enviados à LLM por PDF.")
    p.add_argument("--limit", type=int, default=0)
    p.add_argument("--resume", action="store_true")
    p.add_argument("--write-text", action="store_true", help="Grava textos extraídos em extracted_text/.")
    p.add_argument("--seed", type=int, default=42)
    p.add_argument("--checkpoint-name", default="checkpoint_results.jsonl")
    p.add_argument("--preflight", action="store_true", help="Testa conexão com a LLM com uma chamada curta e encerra.")
    p.add_argument("--no-progressbar", action="store_true", help="Evita tqdm/barra de progresso; melhor para Cursor e logs.")
    p.add_argument("--strict-llm", action="store_true", help="Se ativado, erro de LLM derruba o PDF. Sem isso, registra erro e continua.")
    p.add_argument("--cursor-command", default=os.getenv("CURSOR_AGENT_COMMAND", "agent"), help="Comando do Cursor CLI. O padrão oficial é agent; use cursor-agent apenas se sua instalação expuser esse nome.")
    p.add_argument("--cursor-pass-mode", choices=["stdin", "arg"], default="stdin", help="Como enviar o prompt ao Cursor CLI. stdin é mais seguro para prompts longos.")
    return p.parse_args()


def main() -> None:
    args = parse_args()
    random.seed(args.seed)
    input_path = Path(args.input)
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)
    run_id = f"python_llm_metagrade_{date_id()}"
    pdfs = collect_pdfs(input_path, args.limit)
    checkpoint_path = output_dir / args.checkpoint_name

    if args.provider == "openai" and not args.api_key:
        raise SystemExit("LLM_API_KEY/OPENAI_API_KEY não configurada. Defina a chave ou passe --api-key. Para Cursor Composer, use --provider cursor-cli.")

    # Para Cursor Composer 2.5, o provider é o Cursor CLI. Se o comando não estiver
    # disponível no PATH, falhamos antes de processar PDFs. Isto evita rodar 408 PDFs
    # e produzir um workbook inteiro em erro.
    cursor_cli_preflight_or_exit(args)

    client = LLMClient(
        args.provider,
        args.api_base,
        args.api_key,
        args.model,
        args.timeout,
        args.temperature,
        args.retries,
        cursor_command=args.cursor_command,
        cursor_pass_mode=args.cursor_pass_mode,
    )

    if args.preflight:
        print("Preflight: testando uma chamada curta para a LLM...", flush=True)
        test = client.chat_json("Responda somente JSON válido.", "Retorne exatamente: {\"ok\": true}")
        print(f"Preflight OK: {test}", flush=True)
        return

    processed = {safe_text(r.get("pdf_path")) for r in load_checkpoint(checkpoint_path)} if args.resume else set()
    rows = load_checkpoint(checkpoint_path) if args.resume else []

    manifest: Dict[str, Any] = {
        "run_id": run_id, "script_version": SCRIPT_VERSION, "started_at": now_str(), "input": str(input_path), "output": str(output_dir),
        "pdf_count": len(pdfs), "provider": args.provider, "api_base": args.api_base, "model": args.model, "cursor_command": args.cursor_command, "cursor_pass_mode": args.cursor_pass_mode, "reviewers": args.reviewers,
        "max_chars": args.max_chars, "uses_llm": True, "uses_python_deterministic_triage": True, "project_scope": PROJECT_SCOPE,
        "failures": [], "failures_count": 0,
    }

    print(f"Run ID: {run_id}", flush=True)
    print(f"PDFs encontrados: {len(pdfs)}", flush=True)
    print(f"Modelo: {args.provider} / {args.model}", flush=True)
    print(f"Timeout={args.timeout}s | retries={args.retries} | reviewers={args.reviewers} | max_chars={args.max_chars}", flush=True)
    print(f"Checkpoint: {checkpoint_path}", flush=True)

    iterator = tqdm(list(enumerate(pdfs, start=1)), desc="Processando PDFs com Python + LLM") if (tqdm and not args.no_progressbar) else list(enumerate(pdfs, start=1))
    for idx, pdf in iterator:
        if str(pdf) in processed:
            print(f"[{idx}/{len(pdfs)}] pulando já processado: {pdf.name}", flush=True)
            continue
        print(f"[{idx}/{len(pdfs)}] processando: {pdf.name}", flush=True)
        try:
            row = process_pdf(pdf, output_dir, run_id, client, args)
        except Exception as exc:
            if args.strict_llm:
                raise
            row = {c: "" for c in OUTPUT_COLUMNS}
            row.update({"record_id": "R_" + short_hash(f"ERROR|{pdf}|{now_str()}"), "study_id": "S_" + short_hash(str(pdf)), "run_id": run_id, "script_version": SCRIPT_VERSION, "created_at": now_str(), "raw_file_name": pdf.name, "pdf_path": str(pdf), "audit_status": "error", "data_quality_flags": "processing_or_llm_error", "audit_notes": str(exc), "notes_for_reproducibility": traceback.format_exc()[:3000]})
            manifest["failures"].append({"pdf": str(pdf), "error": str(exc)})
            print(f"    ERRO registrado e execução continua: {exc}", flush=True)
        rows.append(row)
        append_checkpoint(checkpoint_path, row)
        print(f"[{idx}/{len(pdfs)}] checkpoint salvo: {row.get('consensus_final_decision') or row.get('audit_status')}", flush=True)

    manifest["finished_at"] = now_str()
    manifest["failures_count"] = len(manifest["failures"])
    try:
        temp_df = pd.DataFrame(rows)
        manifest["decision_counts"] = temp_df.get("consensus_final_decision", pd.Series(dtype=str)).value_counts().to_dict()
        manifest["needs_attention_count"] = int((temp_df.get("audit_status", "") == "needs_attention").sum()) if "audit_status" in temp_df.columns else 0
    except Exception:
        pass

    write_xlsx(rows, output_dir, manifest)
    print("\nResumo:")
    if rows:
        print(pd.DataFrame(rows).get("consensus_final_decision", pd.Series(dtype=str)).value_counts(dropna=False).to_string())
    print(f"Falhas: {manifest['failures_count']}")


if __name__ == "__main__":
    main()
